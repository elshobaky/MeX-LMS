from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as col_l
import json
import os

MAIN_DIR = os.path.dirname(os.path.abspath(__file__))

books = os.path.join(os.path.join(MAIN_DIR, 'data'), 'books.xlsx')
books_json = os.path.join(MAIN_DIR, 'books.json')

members = os.path.join(os.path.join(MAIN_DIR, 'data'), 'members.xlsx')
members_json = os.path.join(MAIN_DIR, 'members.json')


def convert_to_json(fin, fout):
    js = {}
    out = open(fout, 'w')

    wb = load_workbook(filename=fin)
    for x in wb.get_sheet_names():
        sheet_name = x.encode("UTF-8")
        sheet = wb[x]
        sheet_rows = sheet.max_row
        sheet_cols = sheet.max_column
        sh_js = []
        for r in range(1, sheet_rows + 1):
            cells = ['%s%s' % (col_l(c), r) for c in range(1, sheet_cols + 1)]
            row = [unicode(sheet[cell].value).encode("UTF-8") for cell in cells]
            sh_js.append(row)
        js[sheet_name] = sh_js
        # out.write('%s (rows: %s, cols: %s)\n' % (sheet_name, sheet_rows, sheet_cols))
    out.write(json.dumps(js))

    out.close()


def add_books_to_db(fin=books_json):
    from PyQt4 import QtCore
    from models import Book, Cat
    with open(fin, 'r') as books_file:
        data = books_file.read()
        data = json.loads(data)
    cats = data.keys()
    for c in cats:
        cat = Cat(name=c)
        cat.put()
        cat = Cat.get(name=c)[0]
        cat_id = cat.id
        for book in data[c]:
            title = book[0]
            cat_order = book[1]
            if cat_order.isdigit():
                b = Book.add(title=title, cat_id=cat_id,
                             cat_order=int(cat_order), copies=1,
                             available=True)
                if type(b) in [str, QtCore.QString]:
                    print b, title, cat_order


def add_members_to_db(fin=members_json):
    from PyQt4 import QtCore
    from models import Member
    from assets.validators import valid_email, valid_phone
    with open(fin, 'r') as members_file:
        data = members_file.read()
        data = json.loads(data)
    sheets = data.keys()
    for s in sheets:
        print s
        for member in data[s]:
            if len(member) != 5:
                print 'invalid member'
                return
            mob = member[0]
            mob = mob if valid_phone(mob) else None
            email = member[1]
            email = email if valid_email(email) else None
            grade = member[2]
            name = member[3]
            num = member[4]
            if name and num.isdigit():
                m = Member.add(name=name, mob=mob, email=email, note=grade)
                if type(m) in [str, QtCore.QString]:
                    print m, name, num


def run():
    print 'working on books file'
    print 'converting excel to json ...'
    convert_to_json(fin=books, fout=books_json)
    print 'savind data to db ...'
    add_books_to_db()
    print 'working on members file'
    print 'converting excel to json'
    convert_to_json(fin=members, fout=members_json)
    print 'saving data to db'
    add_members_to_db()


if __name__ == '__main__':
    convert_to_json(fin=members, fout=members_json)
    # add_to_db()
